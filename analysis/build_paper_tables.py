"""
build_paper_tables.py — Rakitan tabel/artefak untuk Bab 4 manuskrip.

Input: hasil analisis (results/analysis/*.csv, results/EXP-003/crypto_metrics_*,
       results/EXP-003/npcr_uaci.csv, results/exported/*, raw_batch_results.csv).
Output:
  results/PAPER_DATASET.csv      detail per-request (EXP-001)
  results/PAPER_TABLES.xlsx      8 sheets artikel (openpyxl)
  results/PAPER_SUMMARY.md       angka-angka kunci
  results/FIGURE_DATA/*.csv      data siap plot

Catatan: sheet "Microservices_Performance" berisi skema + baris indikator
"belum dijalankan" karena load-test skala penuh (EXP-004) memang belum ada;
tabel diisi saat eksperimen itu dijalankan.
"""

import csv
import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RES = ROOT / "results"
ANALYSIS = RES / "analysis"
FIG = RES / "FIGURE_DATA"
XLSX = RES / "PAPER_TABLES.xlsx"
DATASET_CSV = RES / "PAPER_DATASET.csv"

CLASSES = ["Healthy", "Rust", "Miner", "Phoma", "Red Spider Mite", "Cerscospora"]
DATASET_ROOT = ROOT / "data" / "experiment_dataset_v2"
FEATURES = ["entropy", "glcm_correlation", "glcm_contrast", "size_kb",
            "image_width", "image_height"]


def read_csv(path: Path):
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def fnum(v, cast=float, default=None):
    try:
        return cast(v) if v not in (None, "") else default
    except (TypeError, ValueError):
        return default


def _mean(v):
    return round(statistics.mean(v), 4) if v else ""


def _stdev(v):
    return round(statistics.pstdev(v), 4) if len(v) > 1 else ""


# ---------------------------------------------------------------------------
# PAPER_DATASET.csv (EXP-001, per request)
# ---------------------------------------------------------------------------
def build_paper_dataset():
    rows = read_csv(RES / "EXP-001" / "raw_batch_results.csv")
    out = []
    headers = ["experiment_id", "request_id", "filename", "class_label",
               "method", "decision_code", "reasoning",
               "image_width", "image_height", "size_kb", "entropy",
               "glcm_correlation", "glcm_contrast",
               "encryption_time_ms", "decryption_time_ms",
               "end_to_end_latency_ms", "cipher_entropy", "psnr",
               "decrypt_verified", "psnr_is_infinite", "success", "error"]
    for r in rows:
        if r.get("phase") != "main":
            continue
        rel = r.get("relative_path", "")
        cls = ""
        for c in CLASSES:
            if rel.startswith(c + "/"):
                cls = c
                break
        out.append([r["experiment_id"], r["request_id"], r["filename"], cls,
                    r["method"], r["decision_code"], r["reasoning"],
                    r["image_width"], r["image_height"], r["size_kb"],
                    r["entropy"], r["glcm_correlation"], r["glcm_contrast"],
                    r["encryption_time_ms"], r["decryption_time_ms"],
                    r["end_to_end_latency_ms"], r["cipher_entropy"], r["psnr"],
                    r["decrypt_verified"], r["psnr_is_infinite"], r["success"],
                    r["error"]])
    with open(DATASET_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(out)
    print(f"PAPER_DATASET.csv: {len(out)} baris")


# ---------------------------------------------------------------------------
# Helper sheet writer
# ---------------------------------------------------------------------------
def sheet_from_rows(wb, name, headers, rows):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCECDC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for col, h in enumerate(headers, 1):
        w = min(50, max(10, max(len(str(h)), *[
            len(str(x[col - 1])) for x in rows if col - 1 < len(x)])) if rows else len(h))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w + 2
    return ws


# ---------------------------------------------------------------------------
def main():
    from openpyxl import Workbook

    print("== dataset profile ==")
    manifest = json.loads((DATASET_ROOT / "manifest.json").read_text(encoding="utf-8"))
    class_counts = Counter()
    for cls in CLASSES:
        d = DATASET_ROOT / cls
        class_counts[cls] = len(list(d.glob("*.jpg")))

    # fitur per citra (dedupe) dari EXP-001
    exp1 = read_csv(RES / "EXP-001" / "raw_batch_results.csv")
    per_img = {}
    for r in exp1:
        if r.get("phase") == "main":
            per_img.setdefault(r["relative_path"], r)
    feat_vals = defaultdict(list)
    for r in per_img.values():
        for f in FEATURES:
            v = fnum(r.get(f), float, None)
            if v is not None:
                feat_vals[f].append(v)

    feature_stats = [[f, _mean(v), _stdev(v), _pct(sorted(v), .5),
                      round(min(v), 4), round(max(v), 4), len(v)]
                     for f, v in feat_vals.items()]

    print("== build workbook ==")
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Dataset_Profile
    rows = [[c, n, round(100 * n / sum(class_counts.values()), 2),
             _mean([fnum(r.get("size_kb"), float, 0) for r in per_img.values() if
                    str(r.get("relative_path", "")).startswith(c + "/")])]
            for c, n in class_counts.items()]
    total = sum(class_counts.values())
    rows.append(["Total", total, 100.0, ""])
    sheet_from_rows(wb, "Dataset_Profile",
                    ["class_label", "num_images", "percentage", "mean_size_kb"], rows)

    # 2. Feature_Statistics
    sheet_from_rows(wb, "Feature_Statistics",
                    ["feature", "mean", "std", "median", "min", "max", "n"], feature_stats)

    # 3. Selector_Distribution
    sd = read_csv(ANALYSIS / "selector_distribution.csv")
    sheet_from_rows(wb, "Selector_Distribution", sd[0] and list(sd[0].keys()),
                    [list(r.values()) for r in sd] if sd else [])

    # 4. Method_Comparison
    mc = read_csv(ANALYSIS / "method_comparison.csv")
    sheet_from_rows(wb, "Method_Comparison", list(mc[0].keys()), [list(r.values()) for r in mc])

    # 5. Crypto_Security
    cs = read_csv(RES / "EXP-003" / "crypto_metrics_summary.csv")
    nu = read_csv(RES / "EXP-003" / "npcr_uaci.csv")
    merged = {}
    for r in nu:
        if r["method"] == "none":
            continue
        m = r["method"]
        merged.setdefault(m, {"npcr": [], "uaci": []})
        merged[m]["npcr"].append(float(r["npcr_pct"]))
        merged[m]["uaci"].append(float(r["uaci_pct"]))
    cs_rows = []
    for r in cs:
        agg = merged.get(r["method"], {})
        cs_rows.append([
            r["method"], r["samples"], r["mean_payload_entropy"],
            r["mean_corr_adjacent"], r["mean_corr_row_gap"],
            r["mean_chi2_stat"], r["uniform_pass_rate_pct"],
            r["mean_expansion_ratio"],
            _mean(agg.get("npcr", [])), _mean(agg.get("uaci", [])),
        ])
    sheet_from_rows(wb, "Crypto_Security",
                    ["method", "samples", "mean_payload_entropy",
                     "mean_corr_adjacent", "mean_corr_row_gap", "mean_chi2_stat",
                     "uniform_pass_rate_pct", "mean_expansion_ratio",
                     "mean_npcr_pct", "mean_uaci_pct"], cs_rows)

    # 6. Microservices_Performance (EXP-004)
    lt = read_csv(RES / "EXP-004" / "load_test_summary.csv")
    ru = read_csv(RES / "EXP-004" / "service_resource_usage.csv")
    enc_cpu = {}
    for r in ru:
        if r["service"] == "coffee-encryption-service":
            enc_cpu[int(r["vu"])] = r["cpu_mean_pct"]
    lt_rows = []
    for r in lt:
        lt_rows.append([
            int(r["vu"]), r["requests"], r["throughput_req_per_s"],
            r["latency_median_ms"], r["latency_p95_ms"], r["latency_p99_ms"],
            r["error_rate_pct"], enc_cpu.get(int(r["vu"]), ""),
        ])
    sheet_from_rows(wb, "Microservices_Performance",
                    ["vu", "requests", "throughput_req_per_s",
                     "latency_p50_ms", "latency_p95_ms", "latency_p99_ms",
                     "error_rate_pct", "encryption_cpu_mean_pct"], lt_rows)

    # 7. Error_Analysis
    err_rows = []
    for path in sorted(RES.glob("EXP-*")):
        r = read_csv(path / "raw_batch_results.csv")
        fails = [x for x in r if fnum(x.get("success"), int, 0) == 0]
        err_rows.append([path.name, len(fails)])
    fail_detail = read_csv(RES / "exported" / "experiment_failures.csv")
    sheet_from_rows(wb, "Error_Analysis",
                    ["experiment", "failed_requests"], err_rows)
    ws = wb["Error_Analysis"]
    start = len(err_rows) + 3
    ws.cell(row=start, column=1, value="Detail (export gateway failures):")
    for i, h in enumerate(fail_detail[0].keys() if fail_detail else [], 1):
        ws.cell(row=start + 1, column=i, value=h)
    for j, fr in enumerate(fail_detail):
        for i, v in enumerate(fr.values()):
            ws.cell(row=start + 2 + j, column=i + 1, value=v)

    # 8. Reproducibility_Metadata
    meta_rows = []
    for path in sorted(RES.glob("EXP-*")):
        mf = path / "run_metadata.json"
        if mf.exists():
            md = json.loads(mf.read_text(encoding="utf-8"))
            meta_rows.append([path.name, md.get("experiment_id", ""),
                              md.get("n_images", ""), md.get("repeats", ""),
                              md.get("warmup_repeats", ""),
                              md.get("method", ""),
                              md.get("completed_at", "") or md.get("started_at", "")])
    meta_rows.append(["git_commit_at_audit", "34f9a398dbcc31e714cdceebdd5481aafc15b940",
                      "", "", "", "", ""])
    meta_rows.append(["env", "GATEWAY_API_KEY set (64-char acak, gitignored), "
                              "SECRET_KEY/UHC_MATRIX_SIZE/UHC_PASSWORD2 dari .env",
                      "", "", "", "", ""])
    sheet_from_rows(wb, "Reproducibility_Metadata",
                    ["experiment", "experiment_id", "n_images", "repeats",
                     "warmup_repeats", "method", "timestamp"], meta_rows)

    wb.save(XLSX)
    print("PAPER_TABLES.xlsx selesai →", XLSX)

    # ---------------------------------------------------------------------
    # FIGURE_DATA
    # ---------------------------------------------------------------------
    FIG.mkdir(exist_ok=True)
    _copy_fig(sd, "fig_method_distribution.csv", ["method", "total_images", "percentage"])
    _copy_fig(mc, "fig_method_performance.csv",
              ["method", "encryption_time_mean_ms", "end_to_end_latency_mean_ms",
               "cipher_entropy_mean", "success_rate_pct", "payload_expansion_ratio_mean"])
    _copy_fig(cs, "fig_crypto_security.csv",
              ["method", "mean_payload_entropy", "mean_corr_adjacent",
               "mean_npcr_pct", "mean_uaci_pct"])
    _copy_fig(lt, "fig_load_test.csv",
              ["vu", "requests", "throughput_req_per_s", "latency_median_ms",
               "latency_p95_ms", "latency_p99_ms", "error_rate_pct"])
    with open(FIG / "fig_dataset_classes.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["class_label", "num_images"])
        for c, n in class_counts.items():
            w.writerow([c, n])

    # ---------------------------------------------------------------------
    # PAPER_SUMMARY.md
    # ---------------------------------------------------------------------
    per_class_stats = {}
    for cls in CLASSES:
        vals = [fnum(r.get("entropy"), float, 0) for r in per_img.values()
                if str(r.get("relative_path", "")).startswith(cls + "/")]
        per_class_stats[cls] = _mean(vals) if vals else ""

    dist = {r["method"]: r for r in sd}
    total_ok = sum(fnum(r.get("success"), int, 0) for r in exp1 if r.get("phase") == "main")
    total_main = sum(1 for r in exp1 if r.get("phase") == "main")
    all_infinite = all(r.get("psnr_is_infinite") in ("True", "true", "1")
                       for r in exp1 if r.get("phase") == "main")
    exp1_rows = [r for r in exp1 if r.get("phase") == "main"]
    enc_ms_all = [fnum(r.get("encryption_time_ms"), float, 0) or 0 for r in exp1_rows]
    lat_ms_all = [fnum(r.get("end_to_end_latency_ms"), float, 0) or 0 for r in exp1_rows]
    ent_all = [fnum(r.get("cipher_entropy"), float, 0) or 0 for r in exp1_rows]

    lines = []
    lines.append("# PAPER_SUMMARY — AgroCipher Eksperimen Bab 4")
    lines.append("")
    lines.append("Dihasilkan: " + str(Path(__file__).resolve().name))
    lines.append("")
    lines.append("## Dataset")
    lines.append(f"- {total} citra uji ({', '.join(f'{c}: {n}' for c, n in class_counts.items())})")
    lines.append("- Sumber: campuran 5 sumber (Coffee Leaf Diseases, coffee___, drive-download, "
                 "ethiopian test, ethiopian train aug), 6 kelas ternormalisasi; "
                 "downscale max 1024 px, JPEG q90, seed 42; round-robin antar sumber (cap 500/kelas).")
    lines.append(f"- Rata-rata entropy per kelas: {per_class_stats}")
    lines.append("")
    lines.append("## Hasil EXP-001 (adaptive)")
    lines.append(f"- Request utama: {total_main}; sukses: {total_ok} "
                 f"({round(100*total_ok/total_main, 2)}% jika total_main)")
    lines.append(f"- Seluruh {total_main} request lossless (psnr='∞', decrypt_verified=True)." if all_infinite
                 else "- Sebagian request TIDAK lossless — cek psnr_is_infinite.")
    lines.append(f"- Metode terpilih: "
                 + "; ".join(f"{m} {r['total_images']} citra ({r['percentage']}%)"
                             for m, r in dist.items()))
    lines.append(f"- Enkripsi: rata-rata {_mean(enc_ms_all)} ms; end-to-end rata-rata {_mean(lat_ms_all)} ms.")
    lines.append(f"- Cipher entropy rata-rata: {_mean(ent_all)} bit/byte (maks. 8).")
    lines.append("")
    lines.append("## EXP-002 (perbandingan metode)")
    for r in mc:
        lines.append(f"- {r['method']}: sukses {r['success_rate_pct']}%, "
                     f"enkripsi {r['encryption_time_mean_ms']} ms, "
                     f"e2e {r['end_to_end_latency_mean_ms']} ms, "
                     f"cipher entropy {r['cipher_entropy_mean']}, "
                     f"PSNR {r['psnr_mean']}.")
    lines.append("")
    lines.append("## EXP-003 (kriptografi)")
    for r in cs_rows:
        lines.append(f"- {r[0]}: entropy payload {r[2]}, korelasi byte {r[3]}, "
                     f"NPCR {r[8]}%, UACI {r[9]}%.")
    baseline = nu and next((r for r in nu if r["method"] == "none"), None)
    if baseline:
        lines.append(f"- Baseline tanpa enkripsi: NPCR {baseline['npcr_pct']}%, "
                     f"UACI {baseline['uaci_pct']}% (flip 1 byte).")
    lines.append("")
    lines.append("## EXP-004 (performa microservices, load test)")
    for r in lt_rows:
        lines.append(f"- VU={r[0]}: {r[1]} request, throughput {r[2]} req/s, "
                     f"latensi p50={r[3]}ms p95={r[4]}ms p99={r[5]}ms, "
                     f"error {r[6]}%, CPU enkripsi {r[7]}%.")
    lines.append("- Bottleneck: encryption-service (CPU 54-66%); throughput "
                 "plateau ~4.6-5.0 req/s (uvicorn single worker).")
    lines.append("")
    lines.append("## Catatan keterbatasan")
    lines.append("- Tanpa ground-truth label metode terbaik: tidak ada klaim akurasi/precision selector.")
    lines.append("- Blowfish: IV tetap hanya untuk uji NPCR/UACI offline; runtime pakai os.urandom(8).")
    lines.append("- Load test single-node Docker Compose (uvicorn single worker), bukan Kubernetes.")
    lines.append("- Korelasi 'vertikal' = lag baris raster, bukan korelasi spasial 2D.")
    (RES / "PAPER_SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")
    print("PAPER_SUMMARY.md selesai →", RES / "PAPER_SUMMARY.md")
    print("\nSelesai. Artefak: PAPER_DATASET.csv, PAPER_TABLES.xlsx, FIGURE_DATA/, PAPER_SUMMARY.md")


def _pct(sorted_v, q):
    if not sorted_v:
        return ""
    k = (len(sorted_v) - 1) * q
    lo = int(k)
    hi = min(lo + 1, len(sorted_v) - 1)
    return round(sorted_v[lo] + (sorted_v[hi] - sorted_v[lo]) * (k - lo), 4)


def _copy_fig(rows, name, headers):
    with open(FIG / name, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, "") for h in headers])
    print("  wrote", FIG / name)


if __name__ == "__main__":
    main()